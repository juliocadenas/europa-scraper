import os
import logging
import pandas as pd
import csv
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re

logger = logging.getLogger(__name__)

# ==============================================================================
# JOYA DE LA CORONA - COLUMNAS INAMOVIBLES DEL CSV
# ==============================================================================
# Estas columnas son INAMOVIBLES. NO SE PUEDEN CAMBIAR bajo ninguna circunstancia.
# Cualquier cambio aquí rompe toda la integridad del sistema de resultados.
# 
# ESTRUCTURA DEL CSV (INAMOVIBLE):
#   - sic_code:     Código SIC del curso
#   - course_name:  Nombre del curso
#   - title:        Título del resultado encontrado
#   - description:  Descripción del resultado
#   - url:          URL del resultado (UNA SOLA COLUMNA, NO SE DESCOMPONE)
#   - total_words:  Conteo de palabras
#   - lang:         Idioma del resultado (en, es, de, fr, it, pl)
#
# PARA CORDIS MULTI-IDIOMA:
#   - La URL base en inglés se modifica agregando "/" + código de idioma
#   - Ejemplo: https://cordis.europa.eu/article/id/12345/en
#   - Ejemplo: https://cordis.europa.eu/article/id/12345/es
#   - La columna URL contiene la URL completa con el sufijo de idioma
# ==============================================================================
CSV_COLUMNS = [
    'sic_code', 'course_name', 'title', 
    'description', 'url', 'total_words', 'lang'
]
# ==============================================================================

class ResultManager:
    """
    Manages scraping results and storage.
    
    NUEVO MODELO (V3.1):
    - Un archivo CSV por curso: {codigo_sic}-{nombre_curso}.csv
    - Sin carpetas por idioma (todos los resultados en el mismo archivo)
    - El campo 'lang' diferencia los idiomas dentro del mismo archivo
    - Archivo de omitidos: {codigo_sic}-{nombre_curso}_omitidos.xlsx
    
    IMPORTANTE: Las columnas del CSV son INAMOVIBLES. Ver CSV_COLUMNS arriba.
    """
    
    def __init__(self):
        """Initialize the result manager."""
        self.results = []
        self.omitted_results = []
        self.output_file = ""
        self.omitted_file = ""
        self.current_sic_code = ""
        self.current_course_name = ""
    
    def _sanitize_filename(self, name: str) -> str:
        """
        Sanitiza un nombre para usarlo como nombre de archivo.
        Elimina caracteres inválidos y limita la longitud.
        """
        # Reemplazar caracteres inválidos con guión bajo
        sanitized = re.sub(r'[<>:"/\\|?*\[\](){}]', '_', name)
        # Reemplazar múltiples espacios con uno solo
        sanitized = re.sub(r'\s+', '_', sanitized)
        # Limitar longitud
        return sanitized[:50]
    
    def initialize_output_files(self, from_sic: str, to_sic: str, from_course: str, to_course: str, search_engine: str = '', worker_id: Optional[int] = None) -> tuple[str, str]:
        """
        Inicializa los archivos de salida para el RANGO DE CURSOS.
        
        FORMATO DEL NOMBRE DEL ARCHIVO:
        {codigo_desde}-{nombre_desde}_to_{codigo_hasta}-{nombre_hasta}_{timestamp}.csv
        
        Ejemplo: 011901.1-Pea_farms_to_011903.2-Sugar_beets_20260220_165500.csv
        
        Args:
            from_sic: Código SIC inicial del rango
            to_sic: Código SIC final del rango
            from_course: Nombre del curso inicial
            to_course: Nombre del curso final
            search_engine: Motor de búsqueda
            worker_id: ID del worker (para logging)
            
        Returns:
            Tuple of (output_file, omitted_file)
        """
        # Guardar información del rango actual
        self.current_sic_code = from_sic
        self.current_course_name = from_course
        
        # Sanitizar nombres de cursos para el archivo
        sanitized_from_course = self._sanitize_filename(from_course) if from_course else 'unknown'
        sanitized_to_course = self._sanitize_filename(to_course) if to_course else 'unknown'
        
        # Crear timestamp para el nombre del archivo
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # NOMBRE DEL ARCHIVO: {codigo_desde}-{nombre_desde}_to_{codigo_hasta}-{nombre_hasta}_{timestamp}
        base_filename = f"{from_sic}-{sanitized_from_course}_to_{to_sic}-{sanitized_to_course}_{timestamp}"
        
        # Usar rutas absolutas
        project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        results_dir = os.path.join(project_root, 'results')
        os.makedirs(results_dir, exist_ok=True)
        
        # Archivo de resultados: results/{codigo_desde}-{nombre_desde}_to_{codigo_hasta}-{nombre_hasta}_{timestamp}.csv
        self.output_file = os.path.join(results_dir, f"{base_filename}.csv")
        
        # ==============================================================================
        # JOYA DE LA CORONA - CREAR CSV CON COLUMNAS INAMOVIBLES
        # ==============================================================================
        # Usar quoting=csv.QUOTE_ALL para que todos los campos estén entre comillas
        # Esto evita que las comas dentro de los campos se interpreten como delimitadores
        pd.DataFrame(columns=CSV_COLUMNS).to_csv(self.output_file, index=False, quoting=csv.QUOTE_ALL)
        # ==============================================================================
        
        logger.info(f"CSV file created for range {from_sic} to {to_sic}: {self.output_file}")
        
        # Directorio para omitidos
        omitted_dir = os.path.join(results_dir, 'omitidos')
        os.makedirs(omitted_dir, exist_ok=True)
        
        # Archivo de omitidos: results/omitidos/{codigo_desde}-{nombre_desde}_to_{codigo_hasta}-{nombre_hasta}_{timestamp}_omitidos.xlsx
        self.omitted_file = os.path.join(omitted_dir, f"{base_filename}_omitidos.xlsx")
        
        logger.info(f"Omitted file for range {from_sic} to {to_sic}: {self.omitted_file}")
        
        return self.output_file, self.omitted_file
    
    def add_result(self, result: Dict[str, Any]) -> bool:
        """
        Adds a result to the results list and CSV file.
        Todos los resultados van al mismo archivo, diferenciados por el campo 'lang'.
        
        Args:
            result: Result dictionary (debe incluir 'lang')
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Add to results list
            self.results.append(result)
            
            # Add to CSV file (mismo archivo para todos los idiomas)
            return self.append_to_csv(result)
        except Exception as e:
            logger.error(f"Error adding result: {str(e)}")
            return False

    def append_to_csv(self, result: Dict[str, Any]) -> bool:
        """
        Appends a result to the CSV file.
        Todos los resultados van al mismo archivo, el campo 'lang' indica el idioma.
        
        ==============================================================================
        JOYA DE LA CORONA - SOLO SE ESCRIBEN LAS COLUMNAS INAMOVIBLES
        ==============================================================================
        Cualquier campo extra en 'result' que no esté en CSV_COLUMNS será ignorado.
        Esto garantiza que la estructura del CSV sea consistente.
        ==============================================================================
        
        Args:
            result: Result dictionary
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Asegurar que el resultado tenga el campo 'lang'
            if 'lang' not in result:
                result['lang'] = 'en'
            
            # ==============================================================================
            # JOYA DE LA CORONA - FILTRAR SOLO COLUMNAS INAMOVIBLES
            # ==============================================================================
            # Crear un diccionario con SOLO las columnas permitidas
            # Esto evita que campos extra (source, mediatype, rcn, etc.) se agreguen al CSV
            filtered_result = {}
            for col in CSV_COLUMNS:
                filtered_result[col] = result.get(col, '')
            
            df = pd.DataFrame([filtered_result])
            
            # Append sin header (el archivo ya tiene las columnas)
            # Usar quoting=csv.QUOTE_ALL para que todos los campos estén entre comillas
            # Esto evita que las comas dentro de los campos se interpreten como delimitadores
            df.to_csv(self.output_file, mode='a', header=False, index=False, quoting=csv.QUOTE_ALL)
            return True
        except Exception as e:
            logger.error(f"Error appending to CSV: {str(e)}")
            return False
    
    def save_omitted_to_excel(self) -> str:
        """
        Saves omitted results to an Excel file.
        
        Returns:
            Path to the saved file
        """
        try:
            if not self.omitted_results:
                logger.warning("No omitted results to save")
                
                # Even if there are no results, ensure we create an empty Excel file
                if not self.omitted_file:
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    omitted_dir = 'results/omitidos'
                    os.makedirs(omitted_dir, exist_ok=True)
                    self.omitted_file = os.path.join(omitted_dir, f"omitidos_{timestamp}.xlsx")
                
                # Create an empty Excel file
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Omitted Results"
                
                # Add a note that there are no omitted results
                ws.cell(row=1, column=1, value="No se omitieron resultados durante el proceso.")
                
                # Ensure the output directory exists
                output_dir = os.path.dirname(self.omitted_file)
                if output_dir and not os.path.exists(output_dir):
                    os.makedirs(output_dir, exist_ok=True)
                    logger.info(f"Omitted directory created: {output_dir}")
                
                # Save the file
                wb.save(self.omitted_file)
                logger.info(f"Empty omitted results file created at: {self.omitted_file}")
                
                return self.omitted_file
            
            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Omitted Results"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            
            # Add headers
            headers = ['Código SIC', 'Nombre del Curso', 'Título', 'URL', 'Descripción', 'Razón de Omisión']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add data
            for row_num, result in enumerate(self.omitted_results, 2):
                ws.cell(row=row_num, column=1, value=result.get('sic_code', ''))
                ws.cell(row=row_num, column=2, value=result.get('course_name', ''))
                ws.cell(row=row_num, column=3, value=result.get('title', ''))
                ws.cell(row=row_num, column=4, value=result.get('url', ''))
                ws.cell(row=row_num, column=5, value=result.get('description', ''))
                ws.cell(row=row_num, column=6, value=result.get('omission_reason', ''))
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = (max_length + 2) if max_length < 50 else 50
                ws.column_dimensions[column].width = adjusted_width
            
            # Ensure the output directory exists
            output_dir = os.path.dirname(self.omitted_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                logger.info(f"Omitted directory created: {output_dir}")
            
            # Get absolute path
            abs_output_file = os.path.abspath(self.omitted_file)
            
            # Save to Excel
            wb.save(self.omitted_file)
            
            logger.info(f"Saved {len(self.omitted_results)} omitted results to {abs_output_file}")
            
            return abs_output_file
            
        except Exception as e:
            logger.error(f"Error saving omitted results to Excel: {str(e)}")
            return ""
    
    def get_results(self) -> List[Dict[str, Any]]:
        """
        Gets the list of results.
        
        Returns:
            List of result dictionaries
        """
        return self.results
    
    def get_omitted_results(self) -> List[Dict[str, Any]]:
        """
        Gets the list of omitted results.
        
        Returns:
            List of omitted result dictionaries
        """
        return self.omitted_results
    
    def get_output_file(self) -> str:
        """
        Gets the output file path.
        
        Returns:
            Output file path
        """
        return self.output_file
    
    def get_omitted_file(self) -> str:
        """
        Gets the omitted file path.
        
        Returns:
            Omitted file path
        """
        return self.omitted_file
    
    def cleanup_if_empty(self):
        """
        Elimina el archivo CSV si está vacío (solo tiene cabecera).
        """
        try:
            if self.output_file and os.path.exists(self.output_file):
                with open(self.output_file, 'r', encoding='utf-8') as f:
                    lines = f.readlines()
                
                # Si solo tiene la cabecera, eliminar
                if len(lines) <= 1:
                    os.remove(self.output_file)
                    logger.info(f"Removed empty results file: {self.output_file}")
                    
                    # También eliminar el archivo de omitidos si está vacío
                    if self.omitted_file and os.path.exists(self.omitted_file):
                        os.remove(self.omitted_file)
                        logger.info(f"Removed empty omitted file: {self.omitted_file}")
        except Exception as e:
            logger.error(f"Error cleaning up empty files: {e}")
