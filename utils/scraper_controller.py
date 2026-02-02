import asyncio
import logging
import os
import time
import random
from typing import List, Dict, Tuple, Optional, Any, Callable
from datetime import datetime
import re
import pandas as pd

from utils.csv_handler import CSVHandler
from utils.url_processor import URLProcessor
from utils.content_analyzer import ContentAnalyzer
from utils.text_sanitizer import sanitize_filename
from utils.scraper.cordis_api_client import CordisApiClient

# Añadir estas importaciones al inicio del archivo
import os
import sys
import subprocess
import tkinter as tk
from tkinter import messagebox

# Añadir las siguientes importaciones al inicio del archivo (después de las importaciones existentes):
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

class ScraperController:
    """
    Controlador para la aplicación de scraping de USA.gov.
    Gestiona el flujo de trabajo de scraping y procesa resultados.
    """
    
    def __init__(self):
        """Inicializa el controlador de scraping."""
        self.csv_handler = CSVHandler()
        self.url_processor = URLProcessor()
        self.content_analyzer = ContentAnalyzer()
        self.cordis_api_client = CordisApiClient()
        
        self.stop_requested = False
        self.results = []
        self.output_file = ""
        self.processed_urls = set()  # Conjunto para rastrear URLs ya procesadas
        self.stats = {
            'total_urls_found': 0,
            'skipped_non_gov': 0,
            'skipped_low_words': 0,
            'skipped_zero_keywords': 0,
            'skipped_duplicates': 0,
            'saved_records': 0,
            'failed_content_extraction': 0,
            'files_not_saved': 0,
            'total_errors': 0
        }
        
        # Añadir las siguientes variables a la función __init__ (después de self.stats = {...}):
        self.omitted_results = []
        self.omitted_file = ""
        
        # Variables para el progreso basado en resultados
        self.total_results_to_process = 0
        self.processed_results_count = 0
        
        # Variables para rastrear la fase actual
        self.current_phase = 1  # 1 = Búsqueda, 2 = Tabulación
        self.current_tabulation_course = 0  # Contador para la fase de tabulación
    
    async def _check_playwright_browser(self):
        """
        Verifica si el navegador de Playwright está instalado y lo instala si es necesario.
        
        Returns:
            bool: True si el navegador está disponible, False en caso contrario
        """
        try:
            # Verificar si el navegador está instalado buscando en múltiples ubicaciones
            possible_browser_paths = [
                os.path.join(os.path.dirname(os.path.abspath(sys.executable)), 'playwright_browsers'),
                os.path.join(os.getcwd(), 'playwright_browsers'),
                os.path.join(os.path.expanduser('~'), '.cache', 'ms-playwright'),
                os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 'ms-playwright')
            ]
            
            # Establecer la variable de entorno para buscar en estas ubicaciones
            for browser_path in possible_browser_paths:
                if os.path.exists(browser_path):
                    # Buscar específicamente el ejecutable de Chrome
                    chrome_paths = [
                        os.path.join(browser_path, 'chromium-*', 'chrome-win', 'chrome.exe'),
                        os.path.join(browser_path, 'chromium-*', 'chrome-linux', 'chrome'),
                        os.path.join(browser_path, 'chromium-*', 'chrome-mac', 'Chromium.app', 'Contents', 'MacOS', 'Chromium')
                    ]
                    
                    # Usar glob para encontrar coincidencias con el patrón
                    import glob
                    for pattern in chrome_paths:
                        matches = glob.glob(pattern)
                        if matches:
                            logger.info(f"Navegador encontrado en: {matches[0]}")
                            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = browser_path
                            return True
            
            # Si llegamos aquí, intentar lanzar el navegador con la configuración actual
            from playwright.async_api import async_playwright
            
            async with async_playwright() as p:
                try:
                    # Intentar lanzar el navegador
                    browser = await p.chromium.launch()
                    await browser.close()
                    logger.info("Navegador lanzado correctamente")
                    return True
                except Exception as e:
                    logger.warning(f"Error al lanzar el navegador: {str(e)}")
                    
                    # Preguntar al usuario si desea instalar el navegador
                    root = tk.Tk()
                    root.withdraw()  # Ocultar la ventana principal
                    
                    install = messagebox.askyesno(
                        "Navegador no encontrado",
                        "El navegador necesario para el scraping no está instalado. "
                        "¿Desea instalarlo ahora? (Este proceso puede tardar varios minutos)"
                    )
                    
                    if install:
                        try:
                            # Establecer la ruta de los navegadores
                            browsers_path = os.path.join(os.path.dirname(os.path.abspath(sys.executable)), 'playwright_browsers')
                            os.environ['PLAYWRIGHT_BROWSERS_PATH'] = browsers_path
                            
                            # Crear el directorio si no existe
                            os.makedirs(browsers_path, exist_ok=True)
                            
                            # Mostrar mensaje de instalación
                            messagebox.showinfo(
                                "Instalando navegador",
                                "Instalando navegador. Este proceso puede tardar varios minutos. "
                                "La aplicación puede parecer que no responde durante la instalación."
                            )
                            
                            # Instalar el navegador
                            subprocess.run(
                                [sys.executable, '-m', 'playwright', 'install', 'chromium'],
                                check=True
                            )
                            
                            messagebox.showinfo(
                                "Instalación completada",
                                "El navegador se ha instalado correctamente."
                            )
                            
                            return True
                        except Exception as install_error:
                            logger.error(f"Error instalando el navegador: {str(install_error)}")
                            messagebox.showerror(
                                "Error de instalación",
                                f"No se pudo instalar el navegador: {str(install_error)}"
                            )
                            return False
                    else:
                        messagebox.showwarning(
                            "Navegador requerido",
                            "La aplicación requiere el navegador para funcionar correctamente. "
                            "Algunas funciones pueden no estar disponibles."
                        )
                        return False
        except ImportError:
            logger.error("Playwright no está instalado")
            messagebox.showerror(
                "Error",
                "Playwright no está instalado. La aplicación no puede funcionar correctamente."
            )
            return False
    
    def _clean_description(self, text: str) -> str:
        """
        Limpia una descripción eliminando URLs y espacios innecesarios.
        
        Args:
            text: Texto a limpiar
            
        Returns:
            Texto limpio
        """
        if not text:
            return ""
        
        # Eliminar URLs completas (http://, https://, www.)
        cleaned = re.sub(r'https?://\S+', '', text)
        cleaned = re.sub(r'www\.\S+', '', cleaned)
        
        # Eliminar URLs que puedan estar al final del texto
        cleaned = re.sub(r'\s+https?://\S+$', '', cleaned)
        cleaned = re.sub(r'\s+www\.\S+$', '', cleaned)
        
        # Eliminar referencias a URLs como "Visita: " o "Ver en: " al final
        cleaned = re.sub(r'\s+(?:visita|ver|visitar|click|enlace|link|url|ver en|más en|más información en|leer más en)[:\s]+\S+$', '', cleaned, flags=re.IGNORECASE)
        
        # Eliminar cualquier texto que parezca una URL sin protocolo (ejemplo.com)
        cleaned = re.sub(r'\s+\w+\.\w+(?:\.\w+)*(?:/\S*)?', '', cleaned)
        
        # Limpiar espacios múltiples y recortar
        cleaned = re.sub(r'\s+', ' ', cleaned).strip()
        
        return cleaned
    
    # Añadir el siguiente método después del método _clean_description:
    def _save_omitted_to_excel(self, output_file: str) -> str:
        """
        Guarda los resultados omitidos en un archivo Excel.
        
        Args:
            output_file: Ruta para guardar los resultados
            
        Returns:
            Ruta al archivo guardado
        """
        try:
            if not self.omitted_results:
                logger.warning("No hay resultados omitidos para guardar")
                return ""
            
            # Crear un nuevo libro de trabajo
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados Omitidos"
            
            # Definir estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            
            # Añadir encabezados
            headers = ['SIC Code', 'Nombre del Curso', 'Título', 'URL', 'Descripción', 'Razón de Omisión']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Añadir datos
            for row_num, result in enumerate(self.omitted_results, 2):
                ws.cell(row=row_num, column=1, value=result.get('sic_code', ''))
                ws.cell(row=row_num, column=2, value=result.get('course_name', ''))
                ws.cell(row=row_num, column=3, value=result.get('title', ''))
                ws.cell(row=row_num, column=4, value=result.get('url', ''))
                ws.cell(row=row_num, column=5, value=result.get('description', ''))
                ws.cell(row=row_num, column=6, value=result.get('omission_reason', ''))
            
            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                adjusted_width = (max_length + 2) if max_length < 50 else 50
                ws.column_dimensions[column].width = adjusted_width
            
            # Asegurar que el directorio de salida exista
            output_dir = os.path.dirname(output_file)
            if output_dir and not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
                logger.info(f"Directorio de omitidos creado: {output_dir}")
            
            # Obtener ruta absoluta
            abs_output_file = os.path.abspath(output_file)
            
            # Guardar en Excel
            wb.save(output_file)
            
            logger.info(f"Guardados {len(self.omitted_results)} resultados omitidos en {abs_output_file}")
            
            return abs_output_file
            
        except Exception as e:
            logger.error(f"Error guardando resultados omitidos en Excel: {str(e)}")
            return ""
    
    def _should_exclude_result(self, total_words: int, word_counts: Dict[str, int], min_words: int = 30) -> Tuple[bool, str]:
        """
        Determina si un resultado debe ser excluido basado en el conteo de palabras.
        
        Args:
            total_words: Número total de palabras
            word_counts: Diccionario con conteo de cada palabra clave
            min_words: Número mínimo de palabras requeridas
            
        Returns:
            Tupla de (debe_excluir, razón)
        """
        # Reducir el umbral mínimo de palabras para incluir más resultados
        if total_words < min_words:
            return True, f"Total de palabras ({total_words}) menor que el mínimo requerido ({min_words})"
        
        # Ser menos restrictivo con las palabras clave - solo excluir si no hay ninguna coincidencia
        # y hay al menos 3 palabras clave buscadas
        if word_counts and len(word_counts) >= 3 and all(count == 0 for count in word_counts.values()):
            return True, "Ninguna palabra clave encontrada en el contenido"
        
        # No excluir
        return False, ""
    
    def _is_sic_in_range(self, sic_code: str, from_sic: str, to_sic: str) -> bool:
        """
        Verifica si un código SIC está dentro del rango especificado.
        Maneja correctamente códigos SIC con y sin paréntesis.
        
        Args:
            sic_code: Código SIC a verificar
            from_sic: Código SIC inicial del rango
            to_sic: Código SIC final del rango
        
        Returns:
            True si el código está en el rango, False en caso contrario
        """
        try:
            # Si es exactamente igual a alguno de los límites, está en el rango
            if sic_code == from_sic or sic_code == to_sic:
                return True
            
            # Extraer la parte base (antes del paréntesis) y la parte detallada (dentro del paréntesis)
            def extract_parts(code):
                base_match = re.search(r'^(\d+)', code)
                base = int(base_match.group(1)) if base_match else 0
            
                detail_match = re.search(r'$$(\d+)$$', code)
                detail = int(detail_match.group(1)) if detail_match else 0
            
                return base, detail
        
            # Extraer partes de los códigos
            base_code, detail_code = extract_parts(sic_code)
            base_from, detail_from = extract_parts(from_sic)
            base_to, detail_to = extract_parts(to_sic)
        
            # Comparar primero por la parte base
            if base_from < base_code < base_to:
                return True
        
            # Si la base es igual al límite inferior, verificar el detalle
            if base_code == base_from:
                # Si el código a verificar tiene detalle, compararlo con el límite inferior
                if detail_code > 0 and detail_from > 0:
                    return detail_code >= detail_from
                # Si el código a verificar no tiene detalle, está en el rango
                return True
        
            # Si la base es igual al límite superior, verificar el detalle
            if base_code == base_to:
                # Si el código a verificar tiene detalle, compararlo con el límite superior
                if detail_code > 0 and detail_to > 0:
                    return detail_code <= detail_to
                # Si el código a verificar no tiene detalle, está en el rango
                return True
        
            return False
        
        except Exception as e:
            logger.error(f"Error comparando códigos SIC: {str(e)}")
            # En caso de error, usar comparación simple de strings
            return from_sic <= sic_code <= to_sic
    
    async def run_scraping(self, params: Dict[str, Any], progress_callback: Callable = None) -> List[Dict[str, Any]]:
        """
        Ejecuta el proceso de scraping.
        
        Args:
            params: Parámetros de scraping
            progress_callback: Callback para actualizaciones de progreso
            
        Returns:
            Lista de diccionarios de resultados
        """
        try:
            # Extract search engine preference
            search_engine = params.get('search_engine')
            use_api_mode = search_engine == 'Cordis Europa API'
            
            # Verificar si el navegador está disponible (SOLO SI NO ES MODO API)
            if not use_api_mode:
                if not await self._check_playwright_browser():
                    logger.error("Navegador no disponible, no se puede realizar el scraping")
                    if progress_callback:
                        progress_callback(100, "Error: Navegador no disponible", {'total_errors': 1}, 0, 0)
                    return []
            else:
                 logger.info("Modo API activado: Saltando verificación de navegador.")
            # Reiniciar estado
            self.stop_requested = False
            self.results = []
            self.processed_urls = set()
            self.stats = {
                'total_urls_found': 0,
                'skipped_non_gov': 0,
                'skipped_low_words': 0,
                'skipped_zero_keywords': 0,
                'skipped_duplicates': 0,
                'saved_records': 0,
                'failed_content_extraction': 0,
                'files_not_saved': 0,
                'total_errors': 0
            }
            
            # Reiniciar contadores de progreso
            self.total_results_to_process = 0
            self.processed_results_count = 0
            
            # Limpiar la lista de resultados omitidos
            self.omitted_results = []

            # Contadores para cursos
            total_courses = 0
            current_course = 0
            
            # Extraer parámetros
            from_sic = params.get('from_sic')
            to_sic = params.get('to_sic')
            from_course = params.get('from_course', '')
            to_course = params.get('to_course', '')
            min_words = params.get('min_words', 50)
            
            # Validar parámetros
            if not from_sic or not to_sic:
                logger.error("Faltan parámetros requeridos")
                return []
            
            # Registrar los valores de entrada para depuración
            logger.info(f"From Code: {from_sic}, To Code: {to_sic}")
            logger.info(f"From Course: {from_course}, To Course: {to_course}")
            logger.info(f"Mínimo de palabras: {min_words}")
            
            # Cargar los datos del CSV directamente
            self.csv_handler.load_course_data()
            
            # Verificar si es un rango de un solo código SIC
            is_single_sic = from_sic == to_sic
            logger.info(f"¿Rango de un solo código SIC? {is_single_sic}")
            
            # Obtener todos los códigos SIC detallados con cursos
            all_sic_codes_with_courses = self.csv_handler.get_detailed_sic_codes_with_courses()
            
            # Si no se encuentran códigos SIC, registrar y retornar
            if not all_sic_codes_with_courses:
                logger.warning("No se encontraron códigos SIC detallados con cursos en los datos CSV")
                return []
            
            # Filtrar los cursos que están exactamente en el rango especificado
            courses_in_range = []
            
            for sic, course in all_sic_codes_with_courses:
                # Verificar si el código SIC está en el rango especificado
                if self._is_sic_in_range(sic, from_sic, to_sic):
                    courses_in_range.append((sic, course))
                    logger.info(f"Incluido en el rango: {sic} - {course}")
                else:
                    logger.info(f"Excluido del rango: {sic} - {course}")
            
            # Registrar el número de códigos SIC encontrados
            logger.info(f"Encontrados {len(courses_in_range)} códigos SIC en el rango especificado")

            # Establecer el total de cursos
            total_courses = len(courses_in_range)
            logger.info(f"Total de cursos a procesar: {total_courses}")
            
            # Si no se encuentran cursos en el rango, intentar usar coincidencias exactas
            if not courses_in_range:
                logger.warning("No se encontraron cursos en el rango especificado, intentando coincidencias exactas")
                for sic, course in all_sic_codes_with_courses:
                    if sic == from_sic or sic == to_sic:
                        courses_in_range.append((sic, course))
                        logger.info(f"Incluido por coincidencia exacta: {sic} - {course}")
                
                logger.info(f"Encontrados {len(courses_in_range)} cursos usando coincidencias exactas")
            
            # Inicializar progreso
            total_courses = len(courses_in_range)
            current_course_index = 0
            
            if total_courses == 0:
                logger.warning("No se encontraron cursos en el rango especificado")
                return []
            
            # Inicializar contador de archivos encontrados
            total_files_found = 0
            
            # PRIMERA FASE: Recolectar todos los resultados de búsqueda
            all_search_results = []
            
            # Procesar cada curso para recolectar resultados
            for sic_code, course_name in courses_in_range:
                if self.stop_requested:
                    logger.info("Scraping detenido por el usuario")
                    break
              
                current_course += 1
              
                # Mostrar el código y nombre del curso actual en la barra de progreso
                current_course_info = f"{sic_code} - {course_name}"
              
                if progress_callback:
                    # Mantener progreso en 0% durante la fase de búsqueda
                    progress_callback(
                        0,
                        f"Buscando curso {current_course} de {total_courses} - {current_course_info}",
                        self.stats,
                        current_course,
                        total_courses,
                        phase=1  # Indicar explícitamente que estamos en la fase 1
                    )
                
                current_course_index += 1
                
                # Usar el nombre del curso como término de búsqueda
                search_term = course_name
                
                logger.info(f"Buscando '{search_term}' con código {sic_code}")
                
                # Realizar búsqueda con timeout
                try:
                    search_results = []
                    if use_api_mode:
                        # Usar cliente API directo
                        logger.info(f"Usando Cordis API para buscar '{search_term}'")
                        search_results = await self.cordis_api_client.search_projects_and_publications(search_term)
                    else:
                        # Modo web scraping normal con navegador
                        # Establecer un timeout para la búsqueda
                        search_task = self.url_processor.search_cordis_europa(search_term)
                        search_results = await asyncio.wait_for(search_task, timeout=300)  # 5 minutos de timeout
                    
                    # Guardar los resultados con su información de curso
                    for result in search_results:
                        result['sic_code'] = sic_code
                        result['course_name'] = course_name
                        result['search_term'] = search_term
                        all_search_results.append(result)
                    
                    # Actualizar estadísticas
                    self.stats['total_urls_found'] += len(search_results)
                    total_files_found += len(search_results)
                    
                    if progress_callback:
                        progress_callback(
                            0,  # Mantener en 0% durante la búsqueda
                            f"Encontrados {len(search_results)} resultados para '{search_term}' (Total: {len(all_search_results)})",
                            self.stats,
                            current_course,
                            total_courses,
                            phase=1  # Indicar explícitamente que estamos en la fase 1
                        )
                    
                    if not search_results:
                        logger.info(f"No se encontraron resultados para '{search_term}' con código SIC {sic_code}")
                    
                except asyncio.TimeoutError:
                    logger.error(f"Timeout durante la búsqueda de '{search_term}'. Continuando con el siguiente curso.")
                    self.stats['total_errors'] += 1
                    continue
                except Exception as e:
                    logger.error(f"Error durante la búsqueda de '{search_term}': {str(e)}")
                    self.stats['total_errors'] += 1
                    continue
            
            # SEGUNDA FASE: Procesar los resultados recolectados
            
            # Establecer el total de resultados a procesar
            self.total_results_to_process = len(all_search_results)

            # Cambiar a la fase de tabulación
            self.current_phase = 2
            self.current_tabulation_course = 0  # Reiniciar el contador para la fase de tabulación

            if self.total_results_to_process == 0:
                logger.warning("No se encontraron resultados para procesar")
                if progress_callback:
                    progress_callback(100, "No se encontraron resultados para procesar", self.stats, 0, 0)
                return []

            # Notificar explícitamente el cambio a la fase de tabulación con el primer curso
            if progress_callback and len(courses_in_range) > 0:
                first_sic, first_course = courses_in_range[0]
                # Enviar un mensaje explícito de tabulación para el primer curso
                progress_callback(
                    0,  # Iniciar en 0% para la fase de procesamiento
                    f"Tabulando curso 1 de {total_courses} - {first_sic} - {first_course}",
                    self.stats,
                    1,  # Primer curso en tabulación
                    total_courses,
                    phase=2  # Indicar explícitamente que estamos en la fase 2
                )
            
            # Procesar cada resultado
            for i, result in enumerate(all_search_results):
                if self.stop_requested:
                    logger.info("Scraping detenido por el usuario")
                    break
                
                # Actualizar el contador de tabulación
                current_sic_code = result.get('sic_code', '')
                current_course_name = result.get('course_name', '')
                
                # Solo incrementar cuando cambia el curso
                if i == 0 or (current_sic_code != all_search_results[i-1].get('sic_code', '') or 
                             current_course_name != all_search_results[i-1].get('course_name', '')):
                    self.current_tabulation_course += 1
                    
                    # Actualizar progreso con información de tabulación
                    if progress_callback:
                        progress_callback(
                            (self.processed_results_count / self.total_results_to_process) * 100,
                            f"Tabulando curso {self.current_tabulation_course} de {total_courses} - {current_sic_code} - {current_course_name}",
                            self.stats,
                            self.current_tabulation_course,
                            total_courses,
                            phase=2  # Indicar explícitamente que estamos en la fase 2
                        )
                
                # Extraer información del resultado
                sic_code = result.get('sic_code', '')
                course_name = result.get('course_name', '')
                search_term = result.get('search_term', '')
                title = result.get('title', 'Sin Título')
                url = result.get('url', '')
                description = result.get('description', 'Sin Descripción')
                
                # Verificar si la URL ya ha sido procesada
                if url in self.processed_urls:
                    logger.info(f"URL ya procesada, omitiendo: {url}")
                    self.stats['skipped_duplicates'] += 1
                    self.stats['files_not_saved'] += 1
                    
                    # Registrar en resultados omitidos
                    self.omitted_results.append({
                        'sic_code': sic_code,
                        'course_name': course_name,
                        'title': title,
                        'url': url,
                        'description': description,
                        'omission_reason': "URL duplicada"
                    })
                    
                    # Incrementar contador de resultados procesados
                    self.processed_results_count += 1
                    
                    # Actualizar progreso
                    if progress_callback:
                        progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                        progress_callback(
                            progress_percentage,
                            f"Procesando {self.processed_results_count} de {self.total_results_to_process} URL: {url}",
                            self.stats,
                            self.current_tabulation_course,
                            total_courses,
                            phase=2  # Indicar explícitamente que estamos en la fase 2
                        )
                    
                    continue
                
                # Marcar la URL como procesada
                self.processed_urls.add(url)
                
                if not url:
                    # Incrementar contador de resultados procesados
                    self.processed_results_count += 1
                    continue
                
                # Actualizar progreso para cada URL
                if progress_callback:
                    progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                    progress_callback(
                        progress_percentage,
                        f"Procesando {self.processed_results_count+1} de {self.total_results_to_process} URL: {url}",
                        self.stats,
                        self.current_tabulation_course,
                        total_courses,
                        phase=2  # Indicar explícitamente que estamos en la fase 2
                    )
                
                try:
                    # Extraer contenido completo de la URL con timeout
                    logger.info(f"Extrayendo contenido completo de: {url}")
                    
                    try:
                        # Establecer un timeout para la extracción de contenido
                        # Si estamos en modo API, NO permitir uso de navegador
                        allow_browser_extraction = not use_api_mode
                        
                        extract_task = self.url_processor.extract_full_content(url, allow_browser=allow_browser_extraction)
                        full_content = await asyncio.wait_for(extract_task, timeout=180)  # 3 minutos de timeout
                        
                        if not full_content:
                            logger.warning(f"No se pudo extraer contenido de: {url}")
                            self.stats['failed_content_extraction'] += 1
                            self.stats['files_not_saved'] += 1
                            
                            # Registrar en resultados omitidos
                            self.omitted_results.append({
                                'sic_code': sic_code,
                                'course_name': course_name,
                                'title': title,
                                'url': url,
                                'description': description,
                                'omission_reason': "Error en extracción de contenido"
                            })
                            
                            # Incrementar contador de resultados procesados
                            self.processed_results_count += 1
                            
                            # Actualizar progreso
                            if progress_callback:
                                progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                                progress_callback(
                                    progress_percentage,
                                    f"Procesando {self.processed_results_count} de {self.total_results_to_process} URL: {url}",
                                    self.stats,
                                    self.current_tabulation_course,
                                    total_courses,
                                    phase=2  # Indicar explícitamente que estamos en la fase 2
                                )
                            
                            continue
                        
                    except asyncio.TimeoutError:
                        logger.error(f"Timeout durante la extracción de contenido de {url}. Omitiendo URL.")
                        self.stats['failed_content_extraction'] += 1
                        self.stats['total_errors'] += 1
                        self.stats['files_not_saved'] += 1
                        
                        # Registrar en resultados omitidos
                        self.omitted_results.append({
                            'sic_code': sic_code,
                            'course_name': course_name,
                            'title': title,
                            'url': url,
                            'description': description,
                            'omission_reason': "Error en extracción de contenido"
                        })
                        
                        # Incrementar contador de resultados procesados
                        self.processed_results_count += 1
                        
                        # Actualizar progreso
                        if progress_callback:
                            progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                            progress_callback(
                                progress_percentage,
                                f"Procesando {self.processed_results_count} de {self.total_results_to_process} URL: {url}",
                                self.stats,
                                self.current_tabulation_course,
                                total_courses,
                                phase=2  # Indicar explícitamente que estamos en la fase 2
                            )
                        
                        continue
                    except Exception as e:
                        logger.error(f"Error durante la extracción de contenido de {url}: {str(e)}")
                        self.stats['failed_content_extraction'] += 1
                        self.stats['total_errors'] += 1
                        self.stats['files_not_saved'] += 1
                        
                        # Registrar en resultados omitidos
                        self.omitted_results.append({
                            'sic_code': sic_code,
                            'course_name': course_name,
                            'title': title,
                            'url': url,
                            'description': description,
                            'omission_reason': "Error en extracción de contenido"
                        })
                        
                        # Incrementar contador de resultados procesados
                        self.processed_results_count += 1
                        
                        # Actualizar progreso
                        if progress_callback:
                            progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                            progress_callback(
                                progress_percentage,
                                f"Procesando {self.processed_results_count} de {self.total_results_to_process} URL: {url}",
                                self.stats,
                                self.current_tabulation_course,
                                total_courses,
                                phase=2  # Indicar explícitamente que estamos en la fase 2
                            )
                        
                        continue
                    
                    # Contar todas las palabras en el contenido completo
                    total_words = self.url_processor.count_all_words(full_content)
                    
                    # Contar ocurrencias de palabras clave del término de búsqueda
                    word_counts = self.url_processor.estimate_keyword_occurrences(full_content, search_term)
                    
                    # Verificar si el resultado debe ser excluido
                    should_exclude, exclude_reason = self._should_exclude_result(total_words, word_counts, min_words)
                    
                    if should_exclude:
                        if "total de palabras" in exclude_reason.lower():
                            logger.info(f"Excluyendo URL por bajo conteo de palabras: {url} - {exclude_reason}")
                            self.stats['skipped_low_words'] += 1
                            self.stats['files_not_saved'] += 1
                            
                            # Registrar en resultados omitidos
                            self.omitted_results.append({
                                'sic_code': sic_code,
                                'course_name': course_name,
                                'title': title,
                                'url': url,
                                'description': description,
                                'omission_reason': f"Bajo conteo de palabras: {total_words}"
                            })
                        else:
                            logger.info(f"Excluyendo URL por palabras clave con conteo cero: {url} - {exclude_reason}")
                            self.stats['skipped_zero_keywords'] += 1
                            self.stats['files_not_saved'] += 1
                            
                            # Registrar en resultados omitidos
                            self.omitted_results.append({
                                'sic_code': sic_code,
                                'course_name': course_name,
                                'title': title,
                                'url': url,
                                'description': description,
                                'omission_reason': "Sin coincidencias de palabras clave"
                            })
                        
                        # Incrementar contador de resultados procesados
                        self.processed_results_count += 1
                        
                        # Actualizar progreso
                        if progress_callback:
                            progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                            progress_callback(
                                progress_percentage,
                                f"Procesando {self.processed_results_count} de {self.total_results_to_process} URL: {url}",
                                self.stats,
                                self.current_tabulation_course,
                                total_courses,
                                phase=2  # Indicar explícitamente que estamos en la fase 2
                            )
                        
                        continue
                    
                    # Formatear el conteo de palabras
                    formatted_word_counts = self.url_processor.format_word_counts(total_words, word_counts)
                    
                    # Verificar si solo contiene "Total words: X" sin coincidencias de palabras clave
                    if formatted_word_counts.startswith("Total words:") and len(formatted_word_counts.split("|")) == 1:
                        logger.info(f"Excluyendo URL sin coincidencias de palabras clave: {url}")
                        self.stats['skipped_zero_keywords'] += 1
                        self.stats['files_not_saved'] += 1
                        
                        # Registrar en resultados omitidos
                        self.omitted_results.append({
                            'sic_code': sic_code,
                            'course_name': course_name,
                            'title': title,
                            'url': url,
                            'description': description,
                            'omission_reason': "Sin coincidencias de palabras clave"
                        })
                        
                        # Incrementar contador de resultados procesados
                        self.processed_results_count += 1
                        
                        # Actualizar progreso
                        if progress_callback:
                            progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                            progress_callback(
                                progress_percentage,
                                f"Procesando {self.processed_results_count} de {self.total_results_to_process} URL: {url}",
                                self.stats,
                                self.current_tabulation_course,
                                total_courses,
                                phase=2  # Indicar explícitamente que estamos en la fase 2
                            )
                        
                        continue
                    
                    # Limpiar la descripción de URLs y otros elementos no deseados
                    description = self._clean_description(description)

                    # Si la descripción es muy corta, intentar usar parte del contenido completo
                    if len(description) < 100 and full_content:
                        # Tomar los primeros 1000 caracteres del contenido completo
                        content_preview = full_content[:1000]
                        # Limpiar el contenido de URLs y otros elementos no deseados
                        content_preview = self._clean_description(content_preview)
                        
                        # Si el contenido es significativamente más largo que la descripción, usarlo
                        if len(content_preview) > len(description) * 2:
                            description = content_preview
                    
                    # Crear un registro para cada resultado
                    result_data = {
                        'sic_code': sic_code,
                        'course_name': course_name,
                        'title': title,
                        'description': description,
                        'url': url,
                        'total_words': formatted_word_counts
                    }
                    
                    # Añadir a la lista de resultados
                    self.results.append(result_data)
                    
                    # Añadir el resultado al CSV inmediatamente
                    self.csv_handler.append_to_csv(self.output_file, result_data)
                    self.stats['saved_records'] += 1
                    
                    # Incrementar contador de resultados procesados
                    self.processed_results_count += 1
                    
                    # Actualizar progreso
                    if progress_callback:
                        progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                        progress_callback(
                            progress_percentage,
                            f"Procesando {self.processed_results_count} de {self.total_results_to_process} URL: {url}",
                            self.stats,
                            self.current_tabulation_course,
                            total_courses,
                            phase=2  # Indicar explícitamente que estamos en la fase 2
                        )
                    
                    # Pequeña pausa para evitar sobrecarga
                    await asyncio.sleep(random.uniform(0.5, 1))
                    
                except Exception as e:
                    logger.error(f"Error procesando URL {url}: {str(e)}")
                    self.stats['failed_content_extraction'] += 1
                    self.stats['total_errors'] += 1
                    self.stats['files_not_saved'] += 1
                    
                    # Registrar en resultados omitidos
                    self.omitted_results.append({
                        'sic_code': sic_code,
                        'course_name': course_name,
                        'title': title,
                        'url': url,
                        'description': description,
                        'omission_reason': "Error en extracción de contenido"
                    })
                    
                    # Incrementar contador de resultados procesados
                    self.processed_results_count += 1
                    
                    # Actualizar progreso
                    if progress_callback:
                        progress_percentage = (self.processed_results_count / self.total_results_to_process) * 100
                        progress_callback(
                            progress_percentage,
                            f"Procesando {self.processed_results_count} de {self.total_results_to_process} URL: {url}",
                            self.stats,
                            self.current_tabulation_course,
                            total_courses,
                            phase=2  # Indicar explícitamente que estamos en la fase 2
                        )
            
            # Guardar resultados omitidos en Excel
            if self.omitted_results:
                self._save_omitted_to_excel(self.omitted_file)
                logger.info(f"Resultados omitidos guardados en: {os.path.abspath(self.omitted_file)}")
                
                # Actualizar mensaje de progreso final
                if progress_callback:
                    progress_callback(
                        100, 
                        f"Proceso completado. Se encontraron {len(self.results)} resultados y se omitieron {len(self.omitted_results)}.",
                        self.stats
                    )
            
            # Asegurarse de cerrar los recursos
            try:
                await self.url_processor.close()
                await self.content_analyzer.close()
            except Exception as e:
                logger.error(f"Error cerrando recursos: {str(e)}")
                self.stats['total_errors'] += 1
            
            logger.info(f"Resultados guardados en: {os.path.abspath(self.output_file)}")
            logger.info(f"Estadísticas finales: {self.stats}")

            # Actualizar progreso final
            if progress_callback:
                progress_callback(
                    100, 
                    f"Proceso completado. Se encontraron {len(self.results)} resultados.",
                    self.stats
                )

            return self.results
            
        except Exception as e:
            logger.error(f"Error durante el scraping: {str(e)}")
            self.stats['total_errors'] += 1
            
            # Intentar cerrar recursos incluso en caso de error
            try:
                await self.url_processor.close()
                await self.content_analyzer.close()
            except Exception as close_error:
                logger.error(f"Error cerrando navegador después de error: {str(close_error)}")
                self.stats['total_errors'] += 1
            
            if progress_callback:
                progress_callback(100, f"Error: {str(e)}", self.stats, 0, 0)
            
            return []
    
    def append_to_csv(self, result_data):
        """Añade un resultado al archivo CSV."""
        try:
            df = pd.DataFrame([result_data])
            df.to_csv(self.output_file, mode='a', header=False, index=False)
            return True
        except Exception as e:
            logger.error(f"Error añadiendo al CSV: {str(e)}")
            self.stats['total_errors'] += 1
            return False
    
    def update_csv_row(self, result_data):
        """Actualiza una fila en el archivo CSV."""
        try:
            df = pd.read_csv(self.output_file)
            
            # Encontrar la fila coincidente
            mask = (
                (df['sic_code'] == result_data['sic_code']) & 
                (df['title'] == result_data['title']) & 
                (df['url'] == result_data['url'])
            )
            
            if mask.any():
                # Actualizar la fila
                for key, value in result_data.items():
                    if key in df.columns:
                        df.loc[mask, key] = value
                
                # Guardar el DataFrame actualizado
                df.to_csv(self.output_file, index=False)
                
                return True
            else:
                logger.warning(f"No se encontró fila coincidente para actualizar en {self.output_file}")
                return False
            
        except Exception as e:
            logger.error(f"Error actualizando fila CSV: {str(e)}")
            self.stats['total_errors'] += 1
            return False
    
    def stop_scraping(self):
        """Detiene el proceso de scraping."""
        self.stop_requested = True
        logger.info("Detención solicitada por el usuario")
