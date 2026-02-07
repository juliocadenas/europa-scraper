#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
GUI Principal del Scraper
========================
Interfaz gráfica principal para el scraper de USA.gov
"""

import hashlib
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import logging
import sys
import os
import threading
import asyncio
import queue
import re
import subprocess
from datetime import datetime
import time
from pathlib import Path
import webbrowser
import socket
import requests
import json


# Añadir raíz del proyecto al path para permitir imports absolutos
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from gui.styles import setup_styles
from gui.timer_manager import TimerManager
from gui.components.gui_components import (
    ProgressFrame, ResultsFrame, ControlFrame
)
from gui.proxy_config import ProxyConfigWindow
from gui.captcha_config import CaptchaConfigWindow
from gui.config_tab import ConfigTab
from gui.search_config_tab import SearchConfigTab
from utils.logger import setup_logger, get_global_log_handler
from utils.text_sanitizer import sanitize_filename
from utils.proxy_manager import ProxyManager
from utils.config import Config

logger = logging.getLogger(__name__)



class ScraperGUI(ttk.Frame):
    """
    Interfaz gráfica principal para la aplicación de scraping de USA.gov.
    Adaptable a diferentes formatos de datos CSV.
    """

    def __init__(self, master, controller=None):
        """
        Inicializa la GUI del scraper.
        
        Args:
            master: Widget padre
            controller: Instancia del controlador de scraping
        """
        super().__init__(master)
        print("Inicializando ScraperGUI...")
        self.master = master
        self.controller = controller
        print(f"DEBUG: ScraperGUI.__init__, self.controller is {self.controller}")
        self.current_task_id = None
        print("ScraperGUI inicializado")
        
        self.proxy_manager = ProxyManager()
        
        # Determinar la ruta de configuración persistente
        if getattr(sys, 'frozen', False):
            # Si estamos corriendo como ejecutable (PyInstaller), usar el directorio del exe
            base_path = os.path.dirname(sys.executable)
        else:
            # Si estamos en desarrollo, usar la carpeta client del proyecto
            base_path = os.path.join(project_root, 'client')
            
        config_path = os.path.join(base_path, 'config.json')
        logger.info(f"Usando archivo de configuración en: {config_path}")
        self.config = Config(config_path)
        
        self.queue = queue.Queue()
        self.timer_manager = TimerManager(update_callback=self._update_timer_display)
        
        self.results = None
        self.current_course_info = ""
        self.is_updating = False
        
        setup_styles()
        self._setup_ui()
        
        # Confirmar que el botón de carga está visible
        print("✅ Botón 'CARGAR CURSOS (CSV/XLS)' agregado a la pestaña Principal")
        
        # Conectar al servidor y cargar datos iniciales
        self._connect_and_load_initial_data()
        
        global_log_handler = get_global_log_handler()
        global_log_handler.set_callback(self._log_callback)
        
        self.process_queue()
        self.master.bind("<Configure>", self._on_resize)
        self._show_existing_logs()

    def _show_existing_logs(self):
        """Muestra los logs existentes en el área de resultados."""
        try:
            global_log_handler = get_global_log_handler()
            existing_logs = global_log_handler.get_logs()
            
            if existing_logs and hasattr(self, 'results_frame'):
                for log in existing_logs:
                    self.results_frame.add_log(log)
        except Exception as e:
            logger.error(f"Error mostrando logs existentes: {e}")

    def _setup_ui(self):
        """Configura los componentes de la interfaz de usuario."""
        computer_name = socket.gethostname()
        self.master.title(f"Europa Scraper v3 (CON MONITOR) - {computer_name}")
        self.master.geometry("1200x700")
        self.master.minsize(1000, 600)
        self.master.configure(background="#f0f0f0")
        
        self.notebook = ttk.Notebook(self.master)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        self.main_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.main_frame, text="Principal")
        
        self._create_server_config_tab()
        # self._create_task_management_tab() # ELIMINADO
        # self._create_monitor_tab() # ELIMINADO
        
        self.config_tab = ConfigTab(self.notebook, self.config)
        self.search_config_tab = SearchConfigTab(self.notebook, self.config)

        self.paned_window = ttk.PanedWindow(self.main_frame, orient=tk.HORIZONTAL)
        self.paned_window.pack(fill=tk.BOTH, expand=True)
        
        self.left_column = ttk.Frame(self.paned_window)
        self.center_column = ttk.Frame(self.paned_window)
        self.right_column = ttk.Frame(self.paned_window)
        
        self.paned_window.add(self.left_column, weight=4)
        self.paned_window.add(self.center_column, weight=4)
        self.paned_window.add(self.right_column, weight=2)
        
        # --- COLUMNA IZQUIERDA ---
        self.sic_frame = ttk.LabelFrame(self.left_column, text="Selección de Código", padding=10)
        self.sic_frame.pack(fill=tk.BOTH, expand=True, pady=10)
        self.labels_frame = ttk.Frame(self.sic_frame)
        self.labels_frame.pack(fill=tk.X, padx=0, pady=(5, 0))
        ttk.Label(self.labels_frame, text="Curso desde:", font=("TkDefaultFont", 9, "bold")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2.5))
        ttk.Label(self.labels_frame, text="Curso hasta:", font=("TkDefaultFont", 9, "bold")).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2.5, 0))

        self.search_frame = ttk.Frame(self.sic_frame)
        self.search_frame.pack(fill=tk.X, padx=0, pady=5)
        self.search_from_entry = ttk.Entry(self.search_frame)
        self.search_from_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 2.5))
        self.search_from_entry.bind("<KeyRelease>", self._on_search_from)
        self.search_to_entry = ttk.Entry(self.search_frame)
        self.search_to_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(2.5, 0))
        self.search_to_entry.bind("<KeyRelease>", self._on_search_to)
        self.sic_listbox_frame = ttk.Frame(self.sic_frame)
        self.sic_listbox_frame.pack(fill=tk.BOTH, expand=True)
        self.from_sic_listbox = tk.Listbox(self.sic_listbox_frame, exportselection=False, font=("TkDefaultFont", 11)) # Fuente más grande
        self.from_sic_scrollbar = ttk.Scrollbar(self.sic_listbox_frame, orient=tk.VERTICAL, command=self.from_sic_listbox.yview)
        self.from_sic_listbox.configure(yscrollcommand=self.from_sic_scrollbar.set)
        self.from_sic_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.from_sic_scrollbar.pack(side=tk.LEFT, fill=tk.Y)
        self.from_sic_listbox.bind('<<ListboxSelect>>', self._on_from_sic_select)
        self.to_sic_listbox = tk.Listbox(self.sic_listbox_frame, exportselection=False, font=("TkDefaultFont", 11)) # Fuente más grande
        self.to_sic_scrollbar = ttk.Scrollbar(self.sic_listbox_frame, orient=tk.VERTICAL, command=self.to_sic_listbox.yview)
        self.to_sic_listbox.configure(yscrollcommand=self.to_sic_scrollbar.set)
        self.to_sic_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.to_sic_scrollbar.pack(side=tk.LEFT, fill=tk.Y)
        self.to_sic_listbox.bind('<<ListboxSelect>>', self._on_to_sic_select)

        # --- COLUMNA CENTRAL (AHORA MONITOR) ---
        self.header_frame = ttk.Frame(self.center_column)
        self.header_frame.pack(fill=tk.X, pady=(0, 20))
        self.title_label = ttk.Label(self.header_frame, text=f"Europa Scraper - {computer_name}", style="Heading.TLabel")
        self.title_label.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.timer_label = ttk.Label(self.header_frame, text="Tiempo: 00:00:00", style="Timer.TLabel")
        self.timer_label.pack(side=tk.RIGHT, padx=10)
        
        # --- BOTÓN CARGAR CURSOS ---
        self.load_courses_button = ttk.Button(
            self.center_column,
            text="📁 CARGAR CURSOS (CSV/XLS)",
            command=self._upload_courses_file,
            style="Accent.TButton",
            width=50   # Hacer el botón más ancho
        )
        self.load_courses_button.pack(fill=tk.X, pady=10, ipadx=20, ipady=10)  # Más padding para hacerlo más visible
        
        # --- MONITOR DE TAREAS (REEMPLAZA TABLA DE CURSOS) ---
        self.monitor_frame = ttk.LabelFrame(self.center_column, text="Monitor de Progreso en Tiempo Real", padding=10)
        self.monitor_frame.pack(fill=tk.BOTH, expand=True, pady=10)

        # Botones de control del monitor (Footer) - Se empaquetan PRIMERO con side=BOTTOM
        self.monitor_buttons_frame = ttk.Frame(self.monitor_frame)
        self.monitor_buttons_frame.pack(fill=tk.X, pady=5, side=tk.BOTTOM)
        
        self.details_button = ttk.Button(self.monitor_buttons_frame, text="Ver Detalles de Worker", command=self._show_worker_details)
        self.details_button.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True) # EXPANDIDO

        self.force_reset_button = ttk.Button(self.monitor_buttons_frame, text="Forzar Reinicio de Estado", command=self._force_reset_client_state)
        self.force_reset_button.pack(side=tk.RIGHT, padx=5, fill=tk.X, expand=True) # EXPANDIDO
        
        # Treeview para mostrar el estado de los workers (Content)
        # REFACTORIZACIÓN UI: Unificación de Estado y Limpieza de Tarea
        self.worker_tree = ttk.Treeview(self.monitor_frame, columns=('ID', 'Status', 'Course', 'Progress'), show='headings')
        self.worker_tree.heading('ID', text='Worker', command=lambda: self.treeview_sort_column(self.worker_tree, 'ID', False))
        self.worker_tree.heading('Status', text='Estado / Acción', command=lambda: self.treeview_sort_column(self.worker_tree, 'Status', False))
        self.worker_tree.heading('Course', text='Curso / Tarea', command=lambda: self.treeview_sort_column(self.worker_tree, 'Course', False))
        self.worker_tree.heading('Progress', text='Progreso', command=lambda: self.treeview_sort_column(self.worker_tree, 'Progress', False))

        self.worker_tree.column('ID', width=50, anchor=tk.CENTER)
        self.worker_tree.column('Status', width=120)
        self.worker_tree.column('Course', width=350)
        self.worker_tree.column('Progress', width=80, anchor=tk.CENTER)
        
        # Scrollbar para el monitor
        self.monitor_scrollbar = ttk.Scrollbar(self.monitor_frame, orient=tk.VERTICAL, command=self.worker_tree.yview)
        self.worker_tree.configure(yscrollcommand=self.monitor_scrollbar.set)
        
        self.monitor_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.worker_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        # Bind para mostrar detalles al seleccionar
        self.worker_tree.bind('<<TreeviewSelect>>', self._on_tree_select)
        
        # --- MENÚ CONTEXTUAL PARA EL MONITOR ---
        self.monitor_menu = tk.Menu(self.master, tearoff=0)
        self.monitor_menu.add_command(label="Eliminar registro", command=self._delete_selected_records)
        self.worker_tree.bind("<Button-3>", self._show_context_menu) # Clic derecho en Windows

        # --- PANEL DE DETALLES (NUEVO) ---
        self.details_frame = ttk.LabelFrame(self.center_column, text="Detalles del Proceso Seleccionado", padding=5)
        self.details_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.details_text = tk.Text(self.details_frame, height=3, font=("Consolas", 9), state=tk.DISABLED, bg="#f0f0f0")
        self.details_text.pack(fill=tk.X, expand=True)

        # Configuración de motor y controles de inicio (se mantienen igual)
        self.engine_frame = ttk.Frame(self.center_column)
        self.engine_frame.pack(fill=tk.X, pady=5)
        ttk.Label(self.engine_frame, text="Motor de Búsqueda:").pack(side=tk.LEFT, padx=(0, 5))
        self.search_engine_var = tk.StringVar()
        self.search_engine_combo = ttk.Combobox(self.engine_frame, textvariable=self.search_engine_var, values=["Cordis Europa", "Cordis Europa API", "Google", "DuckDuckGo", "Common Crawl", "Wayback Machine"], state="readonly")
        self.search_engine_combo.pack(side=tk.LEFT, fill=tk.X, expand=True)
        self.search_engine_var.set(self.config.get("search_engine", "Cordis Europa"))
        self.control_frame = ControlFrame(self.center_column, on_start=self._on_start_scraping, on_stop=self._on_stop_scraping)
        self.control_frame.pack(fill=tk.X, pady=2)
        self.progress_frame = ProgressFrame(self.center_column)
        self.progress_frame.pack(fill=tk.X, pady=2)
        
        # --- COLUMNA DERECHA ---
        self.results_frame = ResultsFrame(self.right_column)
        self.results_frame.pack(fill=tk.BOTH, expand=True)
        self.results_buttons_frame = ttk.Frame(self.right_column)
        self.results_buttons_frame.pack(fill=tk.X, pady=5)
        self.export_button = ttk.Button(self.results_buttons_frame, text="Exportar Resultados (ZIP)", command=self._on_export_results, state=tk.NORMAL, style="TButton")
        self.export_button.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)
        self.open_folder_button = ttk.Button(self.results_buttons_frame, text="Abrir Carpeta de Resultados", command=self._on_open_results_folder, style="TButton")
        self.open_folder_button.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        # BUTTON CLEANUP: "Limpiar Archivos"
        self.cleanup_button = ttk.Button(
            self.results_buttons_frame, 
            text="Limpiar Archivos", 
            command=self._cleanup_files_action, 
            state=tk.NORMAL
        )
        self.cleanup_button.pack(side=tk.LEFT, padx=5, pady=5, fill=tk.X, expand=True)

        self.paned_window.update()
        width = self.paned_window.winfo_width()
        if width > 0:
            sash_pos1 = int(width * 0.40)
            sash_pos2 = int(width * 0.80)
            self.paned_window.sashpos(0, sash_pos1)
            self.paned_window.sashpos(1, sash_pos2)

    def _create_server_config_tab(self):
        """Crea la pestaña de configuración del servidor"""
        self.server_frame = ttk.Frame(self.notebook, padding="10")
        self.notebook.add(self.server_frame, text="Configuración del Servidor")
        config_frame = ttk.LabelFrame(self.server_frame, text="Configuración", padding=10)
        config_frame.pack(fill=tk.X, pady=10)
        ttk.Label(config_frame, text="URL del Servidor:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.server_url = tk.StringVar(value="https://usr-brief-essex-moss.trycloudflare.com")
        self.server_url_entry = ttk.Entry(config_frame, textvariable=self.server_url, width=50)
        self.server_url_entry.grid(row=0, column=1, padx=10, pady=5, sticky=tk.EW)
        self.connect_button = ttk.Button(config_frame, text="Conectar", command=self._connect_to_server)
        self.connect_button.grid(row=2, column=1, padx=10, pady=20, sticky=tk.E)
        self.connection_status_label = ttk.Label(config_frame, text="Desconectado", foreground="red")
        self.connection_status_label.grid(row=3, column=1, padx=10, sticky=tk.E)
        config_frame.columnconfigure(1, weight=1)

    # _create_task_management_tab y _create_monitor_tab eliminados por refactorización de accesibilidad

    # _create_monitor_tab eliminado por refactorización de accesibilidad

    def _start_status_polling(self):
        """Este método ya no es necesario aquí, la ClientApp se encarga del polling."""
        pass # La ClientApp gestiona ahora el polling y llama a _render_worker_status

    def _on_tree_select(self, event):
        """Muestra los detalles completos del worker seleccionado en el panel inferior."""
        selection = self.worker_tree.selection()
        if not selection:
            return
        
        row_id = selection[0]
        # Recuperar detalles guardados en el mapa
        full_details = getattr(self, 'row_details_map', {}).get(row_id, "Sin detalles disponibles.")
        
        self.details_text.config(state=tk.NORMAL)
        self.details_text.delete(1.0, tk.END)
        self.details_text.insert(tk.END, full_details)
        self.details_text.config(state=tk.DISABLED)

    def _render_worker_status(self, worker_states):
        """Recibe y renderiza el estado detallado de los trabajadores en la GUI."""
        
        # Inicializar mapas de seguimiento si no existen
        if not hasattr(self, 'worker_last_row_id'):
            self.worker_last_row_id = {} 
        if not hasattr(self, 'row_details_map'):
            self.row_details_map = {}

        for worker_id, state in worker_states.items():
            worker_id_str = str(worker_id)
            raw_task = state.get('current_task', 'N/A')
            server_status = state.get('status', 'N/A').capitalize()

            # --- PARSING INTELIGENTE DE STRINGS PARA UI LIMPIA ---
            
            # 1. Determinar STATUS (Acción) y COURSE (Tarea Limpia)
            display_status = server_status # Valor por defecto
            display_course = raw_task      # Valor por defecto
            details_text = raw_task        # Valor por defecto para el panel
            
            # Limpiar estadísticas dinámicas del nombre del curso
            clean_task_text = raw_task.split('|')[0].strip() if '|' in raw_task else raw_task.strip()
            
            if "Iniciando" in raw_task:
                 display_status = "Iniciando"
                 display_course = "Preparando navegador..."
                 is_transient = True
            elif "Esperando" in raw_task:
                 display_status = "Inactivo"
                 display_course = "Esperando tareas..."
                 is_transient = True
            elif "Completado:" in raw_task:
                 display_status = "Completado"
                 # Extract course name from "Completado: X..Y. Total... - COURSE_NAME"
                 if ' - ' in clean_task_text:
                     parts = clean_task_text.split(' - ', 1)
                     display_course = parts[1].strip() if len(parts) > 1 else clean_task_text
                 else:
                     display_course = clean_task_text.replace('Completado:', '').strip()
                 is_transient = False
            elif "Buscando" in raw_task:
                 display_status = "Buscando"
                 # "Buscando curso 1 de 1 - 10.0 - METAL..." -> "10.0 - METAL..."
                 if ' - ' in clean_task_text:
                     parts = clean_task_text.split(' - ', 1)
                     display_course = parts[1].strip() if len(parts) > 1 else clean_task_text
                 is_transient = False
            elif "Tabulando" in raw_task:
                 display_status = "Tabulando"
                 if ' - ' in clean_task_text:
                     parts = clean_task_text.split(' - ', 1)
                     display_course = parts[1].strip() if len(parts) > 1 else clean_task_text
                 is_transient = False
            else:
                 is_transient = False

            
            # 2. Generar ID Único por Tarea (Volver a lógica de historial)
            # El usuario pide EXPRESAMENTE ver cada paso y no solapar.
            
            clean_course_id = display_course.replace(' ', '_').replace('.', '_')
            import re
            clean_course_id = re.sub(r'[^a-zA-Z0-9_]', '', clean_course_id)
            
            # Generar ID basado en el contenido REAL de la tarea y timestamp aproximado (timestamp via worker state?)
            # O mejor: task_scic_code si es posible.
            # Usar clean_course_id es lo mas estable
            
            if is_transient or not clean_course_id or "Esperando" in display_course:
                 # Mensajes de sistema van a fila worker fija
                 current_row_id = f"worker_msg_{worker_id_str}" 
            elif "Completado" in display_status:
                 # Tareas completadas: Usar un ID unico final
                 current_row_id = f"done_{clean_course_id}"
            else:
                 # Tareas activas: ID unico
                 current_row_id = f"active_{clean_course_id}"

            # 3. Formatear Progreso
            progress_val = state.get('progress', 0)
            progress_str = f"{int(progress_val)}%"

            # 4. Datos visuales
            values = (
                worker_id_str,
                display_status,
                display_course,
                progress_str
            )
            
            # 5. Guardar detalles
            full_details_str = raw_task
            if '|' in raw_task:
                 full_details_str = f"Curso: {display_course}\nEstado: {display_status}\n{raw_task.split('|')[1].strip()}"
            
            self.row_details_map[current_row_id] = full_details_str

            # 6. Insertar o Actualizar
            # Lógica Anti-Solapamiento:
            # Si una tarea estaba "Activa" y ahora viene "Completada", borrar la activa y poner la completada
            active_id = f"active_{clean_course_id}"
            done_id = f"done_{clean_course_id}"
            
            if current_row_id == done_id:
                if self.worker_tree.exists(active_id):
                    self.worker_tree.delete(active_id)
                
                # Si ya existe como completada, actualizar (por si llegan msjs repetidos)
                if self.worker_tree.exists(done_id):
                    self.worker_tree.item(done_id, values=values)
                else:
                    self.worker_tree.insert('', 'end', iid=done_id, values=values)
            
            elif current_row_id == active_id:
                # Si esta activa, actualizar
                if self.worker_tree.exists(active_id):
                    self.worker_tree.item(active_id, values=values)
                # Si ya existe como completada (raro per posible), ignorar o borrar completada vieja?
                # Asumimos que si esta activa es lo que vale
                else:
                    self.worker_tree.insert('', '0', iid=active_id, values=values) # Insertar al inicio!

            else:
                # Mensajes transitorios (worker_msg_)
                if self.worker_tree.exists(current_row_id):
                    self.worker_tree.item(current_row_id, values=values)
                else:
                    self.worker_tree.insert('', 'end', iid=current_row_id, values=values)
            
            self.worker_last_row_id[worker_id_str] = current_row_id
            
        # 7. Limpieza Automática de Historial (Max 50 filas completadas)
        all_items = self.worker_tree.get_children()
        completed_items = [item for item in all_items if item.startswith("done_")]
        if len(completed_items) > 50:
            # Borrar las más viejas (las primeras en la lista asumimos)
            for item in completed_items[:-50]:
                self.worker_tree.delete(item)

        self.current_worker_states = worker_states

        # Calcular progreso general promedio
        if worker_states:
            active_workers = [s for s in worker_states.values() if s.get('status') != 'Idle']
            if active_workers:
                total_progress = sum(state.get('progress', 0) for state in active_workers)
                avg_progress = total_progress / len(active_workers)
                self.progress_frame.update_progress(avg_progress, f"Progreso General: {avg_progress:.1f}%")
            else:
                 self.progress_frame.update_progress(100, "Proceso completado")
                 
            # Actualizar estado de botones si hay actividad
            is_working = any(s.get('status') != 'Idle' for s in worker_states.values())
            if is_working:
                self.control_frame.start_button.config(state=tk.DISABLED)
                self.control_frame.stop_button.config(state=tk.NORMAL)
            else:
                self.control_frame.start_button.config(state=tk.NORMAL)
                self.control_frame.stop_button.config(state=tk.DISABLED)
        else:
            self.progress_frame.update_progress(0, "Esperando inicio...")

    def _upload_courses_file(self):
        """Abre el diálogo para seleccionar un archivo y lo sube al servidor."""
        print("DEBUG GUI: _upload_courses_file llamada")
        if self.connection_status_label.cget("text") != "Conectado":
            print("DEBUG GUI: No conectado")
            messagebox.showerror("No Conectado", "Por favor, conéctese a un servidor primero.")
            return

        filepath = filedialog.askopenfilename(
            title="Seleccione un archivo de cursos",
            filetypes=[("Archivos CSV", "*.csv"), ("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
        )
        print(f"DEBUG GUI: Archivo seleccionado: {filepath}")
        if not filepath:
            print("DEBUG GUI: Selección cancelada")
            return

        server_url = self.server_url.get().rstrip('/')
        print(f"DEBUG GUI: URL del servidor: {server_url}")
        self.results_frame.add_log(f"Subiendo archivo {os.path.basename(filepath)} al servidor...")

        try:
            print("DEBUG GUI: Iniciando request POST...")
            with open(filepath, 'rb') as f:
                files = {'file': (os.path.basename(filepath), f)}
                response = requests.post(f"{server_url}/upload_courses", files=files, timeout=30)
            
            print(f"DEBUG GUI: Respuesta recibida. Status: {response.status_code}")
            print(f"DEBUG GUI: Body: {response.text}")

            if response.status_code == 200:
                message = response.json().get("message", "Carga exitosa.")
                messagebox.showinfo("Éxito", message)
                self.results_frame.add_log(message)
                # Refrescar la lista de cursos en la GUI
                self._refresh_courses_from_server()
            else:
                error_detail = response.json().get("detail", response.text)
                messagebox.showerror("Error de Subida", f"El servidor devolvió un error:\n{error_detail}")
                self.results_frame.add_log(f"Error en la subida: {error_detail}")

        except requests.exceptions.RequestException as e:
            print(f"DEBUG GUI: Excepción de request: {e}")
            messagebox.showerror("Error de Red", f"No se pudo comunicar con el servidor: {e}")
            self.results_frame.add_log(f"Error de red: {e}")
        except Exception as e:
            print(f"DEBUG GUI: Excepción general: {e}")
            messagebox.showerror("Error Inesperado", f"Ocurrió un error: {e}")
            self.results_frame.add_log(f"Error inesperado: {e}")

    def _refresh_courses_from_server(self):
        """Obtiene la lista de cursos del servidor y refresca la UI."""
        if self.connection_status_label.cget("text") != "Conectado":
            messagebox.showerror("No Conectado", "Por favor, conéctese a un servidor para refrescar los cursos.")
            return

        server_url = self.server_url.get().rstrip('/')
        self.results_frame.add_log("Refrescando lista de cursos desde el servidor...")

        try:
            response = requests.get(f"{server_url}/get_all_courses", timeout=10)
            if response.status_code == 200:
                courses_data = response.json()
                self._populate_ui_with_course_data(courses_data)
                self.results_frame.add_log(f"Lista de cursos actualizada. {len(courses_data)} cursos cargados.")
            else:
                error_detail = response.json().get("detail", response.text)
                messagebox.showerror("Error de Servidor", f"No se pudo obtener la lista de cursos:\n{error_detail}")
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error de Red", f"No se pudo comunicar con el servidor: {e}")

    def _populate_ui_with_course_data(self, courses_data):
        """Puebla los widgets de la UI con la lista de cursos."""
        try:
            # Convertir a formato consistente (tuplas de strings)
            clean_data = []
            
            for item in courses_data:
                # Manejar tanto diccionarios (servidor) como tuplas (local)
                if isinstance(item, dict):
                    code = str(item.get('sic_code', '')).strip()
                    name = str(item.get('course_name', '')).strip()
                else:
                    code = str(item[0]).strip() if len(item) > 0 else ''
                    name = str(item[1]).strip() if len(item) > 1 else ''
                
                # NO SANITIZAR - Datos tal cual del servidor
                clean_data.append((code, name))
            
            # Guardar datos sin modificar
            self.detailed_sic_codes_with_courses = clean_data
            
            logger.info(f"Cargados {len(self.detailed_sic_codes_with_courses)} cursos para la tabla")

            if not self.detailed_sic_codes_with_courses:
                logger.warning("El servidor devolvió lista vacía de cursos.")
                return
          
            # Limpiar widgets antes de cargar nuevos datos
            self.from_sic_listbox.delete(0, tk.END)
            self.to_sic_listbox.delete(0, tk.END)

            # Poblar la tabla (Treeview) y los Listbox
            for sic_code, course_name in self.detailed_sic_codes_with_courses:
                display_text = f"{sic_code} - {course_name}"
                self.from_sic_listbox.insert(tk.END, display_text)
                self.to_sic_listbox.insert(tk.END, display_text)
          
            logger.info("Datos cargados exitosamente en la tabla y listboxes")
          
        except Exception as e:
            logger.error(f"Error poblando la UI con datos: {str(e)}")
            messagebox.showerror("Error", f"Error mostrando los datos: {str(e)}")

    def _connect_and_load_initial_data(self):
        """Intenta conectar y cargar datos al inicio."""
        self._connect_to_server(show_popups=False)
    
    def _assign_task_to_server(self):
        """Asigna una tarea de scraping al servidor conectado."""
        if self.connection_status_label.cget("text") != "Conectado":
            messagebox.showerror("No Conectado", "Por favor, conéctese a un servidor primero.")
            return

        from_sic = self.from_sic_entry_task.get()
        to_sic = self.to_sic_entry_task.get()

        if not from_sic or not to_sic:
            messagebox.showwarning("Rango no especificado", "Por favor, ingrese los códigos SIC de inicio y fin del trabajo.")
            return

        logger.info(f"Asignando trabajo al servidor: Rango SIC de {from_sic} a {to_sic}")

        try:
            server_url = self.server_url.get().rstrip('/')

            scraping_job_config = {
                'from_sic': from_sic,
                'to_sic': to_sic,
                'min_words': self.config.get('min_words', 30),
                'search_engine': self.search_engine_var.get()
            }

            response = requests.post(f"{server_url}/start_scraping", json=scraping_job_config, timeout=10)

            if response.status_code == 202:
                result = response.json()
                message = result.get('message', 'Trabajo asignado.')
                messagebox.showinfo("Trabajo Asignado", message)
                logger.info(message)
                self.notebook.select(self.monitor_frame) # Cambiar a la pestaña de monitor
            else:
                error_detail = response.json().get("detail", response.text)
                messagebox.showerror("Error de Comunicación", f"Error al asignar trabajo: {error_detail}")

        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error de Red", f"No se pudo conectar con el servidor: {e}")
        except Exception as e:
            messagebox.showerror("Error Inesperado", f"Ocurrió un error inesperado: {e}")


    def _connect_to_server(self, show_popups=True):
        """Conecta al servidor y refresca los datos."""
        try:
            server_url = self.server_url.get().rstrip('/')
            response = requests.get(f"{server_url}/", timeout=5)
            
            if response.status_code == 200:
                self.connection_status_label.config(text="Conectado", foreground="green")
                self.connect_button.config(text="Desconectar", command=self._disconnect_from_server)
                if show_popups:
                    messagebox.showinfo("Conexión Exitosa", "Conectado al servidor correctamente")
                logger.info(f"Conectado al servidor: {server_url}")
                # Cargar datos después de conectar
                self._refresh_courses_from_server()
            else:
                raise Exception(f"Error de conexión: {response.status_code}")
                
        except Exception as e:
            self.connection_status_label.config(text="Desconectado", foreground="red")
            self.connect_button.config(text="Conectar", command=self._connect_to_server)
            if show_popups:
                messagebox.showerror("Error de Conexión", f"No se pudo conectar al servidor:\n{str(e)}")
            logger.error(f"Error conectando al servidor: {str(e)}")

    def _disconnect_from_server(self):
        """Desconecta del servidor"""
        self.connection_status_label.config(text="Desconectado", foreground="red")
        self.connect_button.config(text="Conectar", command=self._connect_to_server)
        messagebox.showinfo("Desconexión", "Desconectado del servidor")
        logger.info("Desconectado del servidor")

    # --- Métodos de Ayuda y Callbacks (sin cambios significativos) ---
    
    def _on_resize(self, event):
        if not self.is_updating:
            self.is_updating = True
            self.is_updating = False

    def _on_from_sic_select(self, event):
        selection = self.from_sic_listbox.curselection()
        if selection:
            selected_text = self.from_sic_listbox.get(selection[0])
            sic_code = selected_text.split(' - ')[0]
            # Actualizar solo si es necesario, por ahora no hay entry que actualizar
            pass

    def _on_to_sic_select(self, event):
        selection = self.to_sic_listbox.curselection()
        if selection:
            selected_text = self.to_sic_listbox.get(selection[0])
            sic_code = selected_text.split(' - ')[0]
            # Actualizar solo si es necesario, por ahora no hay entry que actualizar
            pass

    def _on_search_from(self, event):
        search_term = self.search_from_entry.get().lower()
        self.from_sic_listbox.delete(0, tk.END)
        # Verificar si hay cursos cargados antes de intentar buscar
        if hasattr(self, 'detailed_sic_codes_with_courses') and self.detailed_sic_codes_with_courses:
            for sic_code, course_name in self.detailed_sic_codes_with_courses:
                display_text = f"{sic_code} - {course_name}"
                if search_term in display_text.lower():
                    self.from_sic_listbox.insert(tk.END, display_text)

    def _on_search_to(self, event):
        search_term = self.search_to_entry.get().lower()
        self.to_sic_listbox.delete(0, tk.END)
        # Verificar si hay cursos cargados antes de intentar buscar
        if hasattr(self, 'detailed_sic_codes_with_courses') and self.detailed_sic_codes_with_courses:
            for sic_code, course_name in self.detailed_sic_codes_with_courses:
                display_text = f"{sic_code} - {course_name}"
                if search_term in display_text.lower():
                    self.to_sic_listbox.insert(tk.END, display_text)

    def _update_timer_display(self, time_str):
        self.queue.put(('update_timer', time_str))

    def _on_start_scraping(self):
        """Maneja el clic del botón de inicio de scraping desde la pestaña Principal."""
        # Obtener configuración de todas las pestañas
        try:

            # PRIORIDAD ÚNICA: Usar listas laterales (Accesibilidad)
            from_selection = self.from_sic_listbox.curselection()
            to_selection = self.to_sic_listbox.curselection()
            
            if not from_selection or not to_selection:
                # Si no hay selección, usar cursos cargados o valores por defecto
                if hasattr(self, 'detailed_sic_codes_with_courses') and self.detailed_sic_codes_with_courses:
                    from_sic = self.detailed_sic_codes_with_courses[0][0]  # Primer código
                    to_sic = self.detailed_sic_codes_with_courses[-1][0]  # Último código
                    print(f"🔍 DEPURACIÓN: Usando cursos cargados desde archivo - from_sic='{from_sic}', to_sic='{to_sic}'")
                else:
                    # Último recurso - valores por defecto
                    from_sic = "01.0"
                    to_sic = "011903.0"
                    print(f"🔍 DEPURACIÓN: Usando valores por defecto - from_sic='{from_sic}', to_sic='{to_sic}'")
                    messagebox.showinfo("Selección Automática", f"No se encontraron cursos cargados. Usando rango por defecto:\nDesde: {from_sic}\nHasta: {to_sic}")
            else:
                # Extraer códigos SIC de las listas laterales
                from_sic = self.from_sic_listbox.get(from_selection[0]).split(' - ')[0]
                to_sic = self.to_sic_listbox.get(to_selection[0]).split(' - ')[0]
                
                # DEPURACIÓN: Mostrar los valores extraídos
                print(f"🔍 DEPURACIÓN: from_sic='{from_sic}', to_sic='{to_sic}'")
            
            # Obtener configuración de la pestaña de configuración
            scraping_config = {
                'query': f"{from_sic} a {to_sic}",  # Para compatibilidad con el servidor
                'job_params': {
                    'from_sic': from_sic,
                    'to_sic': to_sic,
                    'min_words': self.config.get('min_words', 30),
                    'num_workers': int(self.config.get('num_workers', 4)), # Enviar número de workers
                    'search_engine': self.search_engine_var.get(),
                    'search_mode': self.config.get('search_mode', 'broad'),
                    'require_keywords': self.config.get('require_keywords', False),
                    'proxy_enabled': self.config.get('proxy_enabled', False),
                    'proxy_rotation': self.config.get('proxy_rotation', False),
                    'captcha_solving_enabled': self.config.get('captcha_solving_enabled', False),
                    'captcha_service': self.config.get('captcha_service', 'manual'),
                    'request_delay': self.config.get('request_delay', 1.0),
                    'page_timeout': self.config.get('page_timeout', 30),
                    'output_format': self.config.get('output_format', 'CSV'),
                    'headless_mode': self.config.get('headless_mode', True)
                }
            }
            
            # Enviar tarea al servidor
            print(f"🔍 DEPURACIÓN FINAL: Enviando configuración: {scraping_config}")
            self._send_scraping_task_to_server(scraping_config)
            
        except Exception as e:
            error_msg = f"Error al iniciar el scraping: {str(e)}"
            messagebox.showerror("Error", error_msg)
            logger.error(f"Error iniciando scraping desde Principal: {str(e)}")
            print(f"❌ ERROR FINAL: {error_msg}")
            # Mostrar el error exacto para depuración
            print(f"❌ ERROR DETALLADO: {error_msg}")
            print(f"🔍 CONFIGURACIÓN ENVIADA: {scraping_config}")
    
    def _send_scraping_task_to_server(self, config):
        """Envía la tarea de scraping al controlador principal (ClientApp)."""
        print(f"DEBUG: _send_scraping_task_to_server, self.controller is {self.controller}")
        if self.connection_status_label.cget("text") != "Conectado":
            messagebox.showerror("No Conectado", "Por favor, conéctese al servidor primero desde la pestaña 'Configuración del Servidor'.")
            return
        
        if not self.controller:
            messagebox.showerror("Error de Arquitectura", "El controlador del cliente no está disponible. La GUI no puede enviar tareas.")
            logger.error("Se intentó enviar una tarea pero self.controller es None.")
            return

        server_url_with_http = self.server_url.get().rstrip('/')
        # El controlador espera la dirección en formato 'host:port' o similar, no la URL completa.
        server_address = server_url_with_http.replace("http://", "").replace("https://", "")
        
        job_params = config.get('job_params', {})
        
        # El controlador se encargará de la lógica de 'is_scraping', logging y el hilo.
        # Ya no se maneja aquí.
        self.controller.start_scraping_on_server(server_address, job_params)
        
        # Cambiamos a la pestaña de monitor para ver el progreso que reportará el controlador.
        # self.notebook.select(self.monitor_frame) # ELIMINADO: Ya está en la pantalla principal

    def _on_stop_scraping(self):
        """Maneja el clic del botón de detención de scraping."""
        server_url = self.server_url.get().rstrip('/')
        try:
            requests.post(f"{server_url}/stop_scraping", timeout=5)
            messagebox.showinfo("Detención Solicitada", "Se ha enviado una solicitud para detener el trabajo al servidor.")
        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error de Red", f"No se pudo enviar la solicitud de detención: {e}")

    def _show_worker_details(self):
        """Muestra una ventana con los detalles de la tarea de un worker seleccionado."""
        selected_item = self.worker_tree.focus() # Obtiene el ID del item seleccionado
        if not selected_item:
            messagebox.showwarning("Selección", "Por favor, seleccione un worker de la lista para ver sus detalles.")
            return

        worker_id = selected_item # 'selected_item' es directamente el iid del item del treeview
        
        if not hasattr(self, 'current_worker_states') or worker_id not in self.current_worker_states:
            messagebox.showerror("Error", f"No se encontraron detalles para el worker ID: {worker_id}")
            return

        details = self.current_worker_states[worker_id]

        detail_window = tk.Toplevel(self.master)
        detail_window.title(f"Detalles del Worker {worker_id}")
        detail_window.transient(self.master) # Hace que la ventana de detalles sea hija de la principal
        detail_window.grab_set() # Captura todos los eventos hasta que se cierre

        ttk.Label(detail_window, text=f"Worker ID: {worker_id}", font=("TkDefaultFont", 10, "bold")).pack(pady=5)
        ttk.Label(detail_window, text=f"Estado: {details.get('status', 'N/A').capitalize()}").pack()
        ttk.Label(detail_window, text=f"Tarea Actual: {details.get('current_task', 'N/A')}").pack()
        ttk.Label(detail_window, text=f"Progreso: {details.get('progress', 0):.2f}%").pack()
        ttk.Label(detail_window, text=f"--- Estadísticas Finales ---", font=("TkDefaultFont", 9, "bold")).pack(pady=5)
        ttk.Label(detail_window, text=f"Procesados: {details.get('processed_count', 0)}").pack()
        ttk.Label(detail_window, text=f"Omitidos: {details.get('omitted_count', 0)}").pack()
        ttk.Label(detail_window, text=f"Errores: {details.get('error_count', 0)}").pack()
        
        ttk.Button(detail_window, text="Cerrar", command=detail_window.destroy).pack(pady=10)

        # Centrar la ventana
        detail_window.update_idletasks()
        x = self.master.winfo_x() + (self.master.winfo_width() // 2) - (detail_window.winfo_width() // 2)
        y = self.master.winfo_y() + (self.master.winfo_height() // 2) - (detail_window.winfo_height() // 2)
        detail_window.geometry(f"+{x}+{y}")

    def _force_reset_client_state(self):
        """Llama al controlador para forzar el reseteo del estado del cliente."""
        if messagebox.askyesno("Confirmar", "¿Está seguro de que desea forzar el reinicio del estado del cliente? Esto debería usarse solo si el cliente se ha quedado atascado en el estado 'en progreso'."):
            if self.controller:
                self.controller.force_reset_state()
                messagebox.showinfo("Reinicio Forzado", "El estado del cliente ha sido reiniciado.")
                
                # LIMPIEZA DE VISTA: Limpiar el monitor al forzar reinicio
                for item in self.worker_tree.get_children():
                    self.worker_tree.delete(item)
                if hasattr(self, 'row_details_map'):
                    self.row_details_map.clear()
                if hasattr(self, 'worker_last_row_id'):
                    self.worker_last_row_id.clear()
            else:
                messagebox.showerror("Error", "El controlador no está disponible para reiniciar el estado.")

    def treeview_sort_column(self, tv, col, reverse):
        """Ordena el contenido de una columna de Treeview."""
        l = [(tv.set(k, col), k) for k in tv.get_children('')]
        
        # Intentar ordenación numérica si es posible (para ID y Progreso)
        try:
            if col in ['ID', 'Progress']:
                def parse_val(val):
                    val = val.replace('%', '').strip()
                    return float(val) if val and val != "N/A" else 0.0
                l = [(parse_val(val), k) for val, k in l]
            else:
                # Ordenación alfabética para el resto (ignorando mayúsculas)
                l = [(val.lower(), k) for val, k in l]
        except (ValueError, TypeError):
            pass

        l.sort(reverse=reverse)

        # Reordenar elementos
        for index, (val, k) in enumerate(l):
            tv.move(k, '', index)

        # Cambiar el comando del encabezado para que la próxima vez sea inverso
        tv.heading(col, command=lambda: self.treeview_sort_column(tv, col, not reverse))

    def _show_context_menu(self, event):
        """Muestra el menú contextual en la posición del ratón."""
        item = self.worker_tree.identify_row(event.y)
        if item:
            # Si el elemento sobre el que se hace clic no está ya seleccionado,
            # cambiar la selección a solo ese elemento.
            # Si YA está seleccionado (parte de una multiselección), no tocamos la selección
            # para no des-sombrear los demás.
            if item not in self.worker_tree.selection():
                self.worker_tree.selection_set(item)
            
            self.monitor_menu.post(event.x_root, event.y_root)

    def _delete_selected_records(self):
        """Elimina los registros seleccionados del monitor."""
        selected_items = self.worker_tree.selection()
        if not selected_items:
            return
            
        if messagebox.askyesno("Confirmar eliminación", f"¿Estás seguro de que deseas eliminar {len(selected_items)} registros de la vista?"):
            for item in selected_items:
                # Limpiar del mapa de detalles si existe
                if hasattr(self, 'row_details_map') and item in self.row_details_map:
                    del self.row_details_map[item]
                
                # Eliminar del treeview
                self.worker_tree.delete(item)
            
            # Limpiar detalles si no queda nada seleccionado
            if not self.worker_tree.selection():
                self.details_text.config(state=tk.NORMAL)
                self.details_text.delete(1.0, tk.END)
                self.details_text.config(state=tk.DISABLED)

    def handle_scraping_finished(self):
        """Maneja el evento de finalización de scraping, actualizando la GUI."""
        self.timer_manager.stop()
        self.control_frame.start_button.config(state=tk.NORMAL)
        self.control_frame.stop_button.config(state=tk.DISABLED)
        self.progress_frame.update_progress(100, "Proceso completado.")

    def _on_export_results(self):
        """Descarga los resultados en un archivo ZIP desde el servidor."""
        if self.connection_status_label.cget("text") != "Conectado":
            messagebox.showerror("No Conectado", "Por favor, conéctese al servidor primero.")
            return

        server_url = self.server_url.get().rstrip('/')
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".zip",
            filetypes=[("Archivos ZIP", "*.zip"), ("Todos los archivos", "*.*")],
            title="Guardar Resultados (ZIP)"
        )
        
        if not filepath:
            return

        self.results_frame.add_log(f"Descargando resultados a: {filepath}...")
        
        try:
            # Petición GET al endpoint de descarga. Usar stream=True para descargas grandes.
            with requests.get(f"{server_url}/download_results", stream=True, timeout=120) as r:
                if r.status_code == 200:
                    with open(filepath, 'wb') as f:
                        for chunk in r.iter_content(chunk_size=8192):
                            f.write(chunk)
                    
                    messagebox.showinfo("Descarga Exitosa", f"Archivo guardado correctamente en:\n{filepath}")
                    self.results_frame.add_log(f"✅ Descarga completada: {filepath}")
                elif r.status_code == 404:
                     messagebox.showwarning("Sin Resultados", "El servidor indica que no hay resultados para descargar.")
                     self.results_frame.add_log("⚠️ El servidor no tiene resultados.")
                else:
                    error_detail = r.json().get("detail", r.text) if r.headers.get('content-type') == 'application/json' else r.text
                    messagebox.showerror("Error de Descarga", f"El servidor devolvió un error:\n{error_detail}")
                    self.results_frame.add_log(f"❌ Error en la descarga: {error_detail}")

        except requests.exceptions.RequestException as e:
            messagebox.showerror("Error de Red", f"No se pudo descargar el archivo: {e}")
            self.results_frame.add_log(f"❌ Error de red durante descarga: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error inesperado: {e}")
            self.results_frame.add_log(f"❌ Error inesperado: {e}")


    def _cleanup_files_action(self):
        """Llama al endpoint /cleanup_files."""
        if messagebox.askyesno("Confirmar Limpieza", "¿Estás seguro? Esto borrará TODOS los resultados y archivos omitidos en el servidor."):
            try:
                server_url = self.server_url.get().rstrip('/')
                response = requests.get(f"{server_url}/cleanup_files", timeout=10)
                if response.status_code == 200:
                    data = response.json()
                    messagebox.showinfo("Limpieza Exitosa", data.get("message", "Limpieza ok"))
                    self.add_log("Limpieza de archivos solicitada y completada.")
                else:
                    messagebox.showerror("Error", f"Error en limpieza: {response.status_code}")
            except Exception as e:
                messagebox.showerror("Error", f"Fallo al conectar para limpieza: {e}")

    def _on_open_results_folder(self):
        results_dir = os.path.abspath("results")
        if not os.path.exists(results_dir):
            os.makedirs(results_dir, exist_ok=True)
        try:
            if os.name == 'nt': os.startfile(results_dir)
            elif os.name == 'posix': subprocess.run(['xdg-open', results_dir])
        except Exception as e:
            logger.error(f"Error abriendo carpeta de resultados: {str(e)}")

    def _load_courses_from_file(self):
        """Carga cursos desde un archivo CSV o XLS con 2 columnas"""
        try:
            filepath = filedialog.askopenfilename(
                title="Seleccionar archivo de cursos",
                filetypes=[
                    ("Archivos Excel", "*.xlsx *.xls"),
                    ("Archivos CSV", "*.csv"),
                    ("Todos los archivos", "*.*")
                ]
            )
            
            if not filepath:
                return
            
            self.results_frame.add_log(f"Cargando cursos desde: {os.path.basename(filepath)}")
            
            # Procesar según el tipo de archivo
            # FIX DEFINITIVO: Usar openpyxl directamente para leer celdas como TEXTO
            if filepath.endswith(('.xlsx', '.xls')):
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(filepath, data_only=True)
                    ws = wb.active
                    
                    courses_data = []
                    for row in ws.iter_rows(min_row=1, values_only=False):
                        if len(row) >= 2:
                            # Obtener el valor de la celda como STRING PURO
                            cell_code = row[0]
                            cell_name = row[1]
                            
                            # Obtener valor como está almacenado en Excel
                            code_val = cell_code.value
                            name_val = cell_name.value
                            
                            if code_val is not None and name_val is not None:
                                # Convertir a string exactamente como aparece
                                code_str = str(code_val).strip()
                                name_str = str(name_val).strip()
                                
                                # Saltar si es la fila de encabezado
                                if code_str.lower() not in ['code', 'sic', 'sic_code', 'curso']:
                                    courses_data.append((code_str, name_str))
                    
                    wb.close()
                except ImportError:
                    messagebox.showerror("Error", "Para archivos Excel necesita instalar openpyxl:\\npip install openpyxl")
                    return
            else:
                # CSV - El formato es: id, sic_code, course_name
                # Necesitamos columna 1 (sic_code) y columna 2 (course_name)
                import csv
                with open(filepath, 'r', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    courses_data = []
                    for row in reader:
                        # Saltar header si existe
                        if len(row) >= 3 and row[1] and row[2]:
                            # row[1] = sic_code, row[2] = course_name
                            if row[1].lower() not in ['sic_code', 'code', 'sic']:
                                courses_data.append((str(row[1]), str(row[2])))
            
            # La sanitización ahora se hace en _update_ui_with_loaded_courses

            if not courses_data:
                messagebox.showwarning("Archivo Vacío", "No se encontraron cursos válidos en el archivo.")
                return
            
            # Actualizar la interfaz con los nuevos cursos
            self._update_ui_with_loaded_courses(courses_data)
            
            self.results_frame.add_log(f"✅ Cargados {len(courses_data)} cursos exitosamente")
            messagebox.showinfo("Éxito", f"Se cargaron {len(courses_data)} cursos desde el archivo.")
            
        except ImportError as e:
            if 'pandas' in str(e):
                messagebox.showerror("Error", "Para archivos Excel necesita instalar pandas:\npip install pandas openpyxl")
            else:
                messagebox.showerror("Error", f"Falta una librería: {e}")
        except Exception as e:
            messagebox.showerror("Error", f"Error al cargar el archivo: {str(e)}")
            self.results_frame.add_log(f"❌ Error cargando archivo: {str(e)}")
    
    def _update_ui_with_loaded_courses(self, courses_data):
        """Actualiza la UI con los cursos cargados desde archivo o servidor"""
        try:
            # Convertir a formato consistente (tuplas de strings)
            # NO SANITIZAR NADA - Mostrar datos exactamente como vienen del CSV
            clean_data = []
            for code, name in courses_data:
                clean_data.append((str(code).strip(), str(name).strip()))
            
            courses_data = clean_data

            # Limpiar widgets existentes
            self.course_tree.delete(*self.course_tree.get_children())
            self.from_sic_listbox.delete(0, tk.END)
            self.to_sic_listbox.delete(0, tk.END)
            
            # Guardar datos para búsquedas
            self.detailed_sic_codes_with_courses = courses_data
            
            # Poblar la tabla y listas
            # FIX CRÍTICO: Tkinter Treeview convierte "01.0" a 1.0 automáticamente
            # Solución: Añadir un espacio de ancho cero (U+200B) al inicio para forzar string
            for sic_code, course_name in courses_data:
                # Añadir zero-width space para prevenir conversión numérica
                display_code = '\u200b' + sic_code  # U+200B = zero-width space
                self.course_tree.insert("", tk.END, values=(display_code, course_name))
                display_text = f"{sic_code} - {course_name}"  # En listbox NO necesita el fix
                self.from_sic_listbox.insert(tk.END, display_text)
                self.to_sic_listbox.insert(tk.END, display_text)
            
            logger.info(f"UI actualizada con {len(courses_data)} cursos desde archivo")
            
        except Exception as e:
            logger.error(f"Error actualizando UI con cursos cargados: {e}")
            messagebox.showerror("Error", f"Error actualizando la interfaz: {str(e)}")

    def _log_callback(self, log_entry):
        self.queue.put(('log', log_entry))

    def process_queue(self):
        try:
            while not self.queue.empty():
                message_type, message = self.queue.get_nowait()
                if message_type == 'update_timer':
                    if hasattr(self, 'timer_label'): self.timer_label.config(text=message)
                elif message_type == 'log':
                    if hasattr(self, 'results_frame'): self.results_frame.add_log(message)
        except queue.Empty:
            pass
        finally:
            self.master.after(200, self.process_queue)

if __name__ == "__main__":
    # Este bloque ahora asegura que la aplicación se inicie correctamente,
    # incluso si este script se ejecuta directamente.
    try:
        # Importar aquí para evitar dependencias circulares en el ámbito global
        from client.main import ClientApp
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
        app = ClientApp()
        app.run()
    except Exception as e:
        try:
            messagebox.showerror("Error Crítico", f"Error al iniciar la aplicación:\n{str(e)}")
        except:
            print(f"Error crítico: {e}")
